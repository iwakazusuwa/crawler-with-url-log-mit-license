# ---
# jupyter:
#   jupytext:
#     formats: py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.17.1
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %%

import datetime
import os
import os.path as osp
from icrawler.builtin import BingImageCrawler
from icrawler.builtin import GoogleImageCrawler
from icrawler import ImageDownloader
from selenium.webdriver.remote.remote_connection import LOGGER as selenium_logger 
import csv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from PIL import Image as PILImage


from logging import getLogger, StreamHandler, DEBUG
logger = getLogger(__name__)
handler = StreamHandler()
handler.setLevel(DEBUG)
logger.setLevel(DEBUG)
logger.addHandler(handler)
logger.propagate = False



# %%
# 検索ワード

a='ペンギン 親子 かわいい'
b='子猫 へそ天 いやし'
c='かき氷 フルーツ エモイ'

keywords = [a, b, c]


# 取得する画像数
kazu=5

# --- JSTタイムゾーンで現在時刻を取得 ---
t_delta = datetime.timedelta(hours=9)
JST=datetime.timezone(t_delta, 'JST')
now = datetime.datetime.now(JST)
day_foi = format(now, '%Y%m%d9%H%M%S')

# --- 保存用ディレクトリ作成 ---
foi = './image_' + str(day_foi)
os.makedirs(foi, exist_ok=True) 

# --- URLリストのファイル名 ---
save_name = 'リスト.csv' 

# --- URLリストに画像を張り付けたエクセルファイル名 ---
save_Exname = 'output.xlsx'


# %%

# %%
class URLDownloader(ImageDownloader):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        import logging
        self.logger.setLevel(logging.CRITICAL)    
    
    def save_column(self, folname, filepath, file_url, output_csv_path=None):
        if output_csv_path is None:
            output_csv_path = os.path.join(foi, save_name)
        with open(output_csv_path, 'a', encoding='utf-8-sig', newline='') as f:
            output_str = f'{folname}, {filepath}, {file_url}\n'
            f.write(output_str)

    def download(self, task, default_ext, timeout=5, max_retry=3, overwrite=False, **kwargs):
        file_url = task['file_url']
        task['success'] = False
        task['filename'] = None
        retry = max_retry

        while retry > 0 and not self.signal.get('reach_max_num'):
            try:
                if not overwrite:
                    with self.lock:
                        self.fetched_num += 1
                        filename = self.get_filename(task, default_ext)
                        if self.storage.exists(filename):
                            self.logger.info('skip downloading file %s', filename)
                            return
                        self.fetched_num -= 1

                response = self.session.get(file_url, timeout=timeout)

                if self.reach_max_num():
                    self.signal['reach_max_num'] = True
                    break

                if response.status_code != 200:
                    self.logger.error('Response status code %d, file %s',
                                      response.status_code, file_url)
                    break

                if not self.keep_file(task, response, **kwargs):
                    break

                with self.lock:
                    self.fetched_num += 1
                    filename = self.get_filename(task, default_ext)

                self.logger.info('image #%s\t%s', self.fetched_num, file_url)
                self.storage.write(filename, response.content)
                task['success'] = True
                task['filename'] = filename
                #folname = task.get('folname', 'unknown')
                self.save_column(folname, filename, file_url)
                break

            except Exception as e:
                self.logger.error('Exception caught when downloading file %s, error: %s, remaining retry times: %d',
                                  file_url, e, retry - 1)
            finally:
                retry -= 1

    def get_filename(self, task, default_ext):
        filename = f"{self.fetched_num}.{default_ext}"
        return filename



# %%

crawler = BingImageCrawler(storage ={'root_dir' : foi})

for keyword in keywords:
    if keyword == a:
        moji = '1_'
    elif keyword == b:
        moji = '2_'
    elif keyword == c:
        moji = '3_'
    # 検索ワードの頭に連番。VBAで使用するので    
    folname = moji + keyword
    
    crawler = BingImageCrawler(downloader_cls=URLDownloader, storage={'root_dir': foi + '/'+ folname})   
    crawler.crawl(keyword = keyword, max_num = kazu) 

# %%
# 画像保存トップフォルダ
base_dir = foi  # ダウンロード時のフォルダ変数と同じにする

# CSVファイルパス
csv_path = os.path.join(foi, 'リスト.csv')

wb = Workbook()
ws = wb.active

# セルの高さ・幅設定
row_height = 84
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 50

# 画像リサイズの高さ
resize_height = row_height  

for i, row in enumerate(csv.reader(open(csv_path, newline='', encoding='utf-8-sig')), start=1):
    folname = row[0].strip()
    img_filename = row[1].strip()
    img_url = row[2].strip()

    img_path = os.path.join(base_dir, folname, img_filename)

    if not os.path.exists(img_path):
        print(f"画像ファイルが見つかりません: {img_path}")
        continue

    # 行の高さ設定（画像サイズに合わせて）
    ws.row_dimensions[i].height = resize_height * 0.75  # Excelの高さはポイントなので調整

    # 画像をPillowで開いてリサイズ（縦resize_heightピクセルに）
    pil_img = PILImage.open(img_path)
    w, h = pil_img.size
    new_h = resize_height
    new_w = int(w * (new_h / h))
    pil_img = pil_img.resize((new_w, new_h))

    # 一時ファイルに保存
    tmp_path = os.path.join(base_dir, f'tmp_resized_{i}.png')
    pil_img.save(tmp_path)

    # Excelに画像挿入
    img = ExcelImage(tmp_path)
    img.anchor = f'A{i}'
    ws.add_image(img)

    # B〜D列にテキスト挿入
    ws[f'B{i}'] = folname
    ws[f'C{i}'] = img_filename

    # URLはハイパーリンク付きテキストで中央揃えに
    cell = ws[f'D{i}']
    cell.value = img_url
    cell.hyperlink = img_url
    cell.style = "Hyperlink"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # B, C列も中央揃え
    for col_letter in ['B', 'C']:
        ws[f'{col_letter}{i}'].alignment = Alignment(horizontal='center', vertical='center')

# 保存して一時ファイルは任意で削除
output_xlsx_path = os.path.join(foi, save_Exname)
wb.save(output_xlsx_path)
print(f"Excelファイルを保存しました: {output_xlsx_path}")

# 一時画像削除（必要なら）
for i in range(1, ws.max_row + 1):
    tmp_file = os.path.join(base_dir, f'tmp_resized_{i}.png')
    if os.path.exists(tmp_file):
        os.remove(tmp_file)


os.startfile(os.path.realpath(foi))
os.startfile(os.path.join(os.path.realpath(foi), save_Exname))
